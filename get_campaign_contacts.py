import pandas as pd
import os
import requests
import json
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# Base URL for the API
base_url = "https://api.givebutter.com/v1/"

# Headers with authorization and accepting JSON response
headers = {
    "accept": "application/json",
    "Authorization": f"Bearer {os.getenv('GIVEBUTTER_API_TOKEN')}",
}
auth_token = os.getenv("GIVEBUTTER_API_TOKEN")


def get_campaign(auth_token):
    """
    Function to retrieve the ID of a specific campaign.

    Parameters:
    auth_token (str): Authorization token for API access.
    """
    global campaign_id
    url = "https://api.givebutter.com/v1/campaigns"
    response = requests.get(url, headers=headers)

    if response.status_code == 200:
        try:
            # Parse the text response as JSON
            campaign_data = json.loads(response.text)

            # Extract the campaign ID from the nested structure
            if "data" in campaign_data and len(campaign_data["data"]) > 0:
                campaign_id = campaign_data["data"][0]["id"]
                print("Campaign ID:", campaign_id)
                return campaign_id
            else:
                print("No campaign data found in response.")
        except json.JSONDecodeError:
            print("Error parsing response as JSON. Response text:", response.text)
    else:
        print(f"Failed to retrieve campaign: Status code {response.status_code}")
        print("Response:", response.text)


def get_campaign_members(campaign_id):
    """
    Function to retrieve members of a specific campaign using its ID.

    Parameters:
    campaign_id (int): The ID of the campaign.
    auth_token (str): Authorization token for API access.

    Returns:
    str: Text response from the API call, which contains the members' data.
    """
    page = 1
    all_members = []
    while True:
        members_url = (
            f"https://api.givebutter.com/v1/campaigns/{campaign_id}/members?page={page}"
        )
        response = requests.get(members_url, headers=headers)

        if response.status_code == 200:
            # Parse the response as JSON
            try:
                members_data = json.loads(response.text)
            except json.JSONDecodeError:
                return f"Failed to parse response as JSON: {response.text}"

            # Extract the member data from the response
            member_data = members_data["data"]
            all_members.extend(member_data)
            page += 1
            # Check if there are more pages
            if (
                members_data["meta"]["current_page"]
                <= members_data["meta"]["last_page"]
            ):
                continue
            else:
                break
        else:
            return f"Failed to retrieve members: Status code {response.status_code}\nResponse: {response.text}"
    # Print the member data
    df2 = pd.DataFrame(all_members)
    return df2


def get_tickets():
    page = 1
    all_tickets = []

    while True:
        ticket_url = f"https://api.givebutter.com/v1/tickets?page={page}"
        response = requests.get(ticket_url, headers=headers)

        if response.status_code == 200:
            # Parse the response as JSON
            tickets_data = json.loads(response.text)

            # Extract the ticket data from the response
            ticket_data = tickets_data["data"]

            # Add the current page's tickets to all_tickets
            all_tickets.extend(ticket_data)

            # Increment the page number
            page += 1

            # Check if there are more pages
            if (
                tickets_data["meta"]["current_page"]
                <= tickets_data["meta"]["last_page"]
            ):
                continue
            else:
                break
        else:
            return f"Failed to retrieve tickets: Status code {response.status_code}\nResponse: {response.text}"
    return all_tickets


def format_data():
    df = pd.DataFrame(get_tickets())

    # Convert email addresses to lowercase
    df["email"] = df["email"].str.lower()

    # Extract the desired columns
    columns = [
        "name",
        "first_name",
        "last_name",
        "email",
        "phone",
        "title",
        "price",
        "created_at",
    ]

    data = df[columns].copy()

    # Convert 'created_at' to datetime and format it to 'YYYY-MM-DD'
    data["created_at"] = pd.to_datetime(data["created_at"]).dt.strftime("%Y-%m-%d")
    data = data.rename(
        columns={
            "name": "Name",
            "first_name": "First",
            "email": "Email",
            "phone": "Phone",
            "title": "Title",
            "price": "Price",
            "created_at": "Signup Date",
        }
    )

    # Group the data by title and get the count of each title
    title_counts = data.groupby("Title").size()

    # Create a new Excel workbook and add a worksheet for each title
    wb = openpyxl.Workbook()
    # Remove the default sheet created
    default_sheet = wb["Sheet"]
    wb.remove(default_sheet)

    for title, count in title_counts.items():
        # Split the sheet name on "-" and take only the part after the last "-"
        sheet_name = title.rsplit("-", 1)[-1]
        # Replace any "/" characters in the sheet name with "_"
        sheet_name = sheet_name.replace("/", "_")
        ws = wb.create_sheet(sheet_name)
        # Add the ticket data to the worksheet
        ws.append(data.columns.tolist())  # Append the updated column names
        filtered_data = data[data["Title"] == title]
        for index, row in filtered_data.iterrows():
            ws.append(row.tolist())

        # Auto-adjust column width based on data
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

    # Save the workbook to an Excel file
    file_name = "GiveButterReport.xlsx"
    path = os.path.join(os.getcwd(), file_name)
    wb.save(path)
    print(f"Generated Excel file: {path}")


def fundraising(df2: pd.DataFrame, file_name: str = "Fundraising_Progress.xlsx"):
    # Drop the 'picture', 'items', and 'url' columns
    df2 = df2.drop(columns=["id", "picture", "items", "url"])

    # Rename the columns
    df2 = df2.rename(
        columns={
            "first_name": "First Name",
            "last_name": "Last Name",
            "display_name": "Display Name",
            "email": "Email",
            "phone": "Phone",
            "raised": "Raised",
            "goal": "Goal",
            "donors": "Donors",
            # Add more columns here if needed
        }
    )

    # Calculate the total raised
    total_raised = df2["Raised"].sum()

    # Create a new Excel workbook
    wb = openpyxl.Workbook()

    # Remove the default sheet created
    default_sheet = wb["Sheet"]
    wb.remove(default_sheet)

    # Create a new worksheet for df2
    ws = wb.create_sheet("Fundraising")

    # Write the data from df2 to the worksheet
    for r in dataframe_to_rows(df2, index=False, header=True):
        ws.append(r)

    # Write the total raised to the bottom of the 'Raised' column
    ws.append(["", "", "", "", "Total Raised", total_raised, "", ""])

    # Auto-adjust column width based on data
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook
    wb.save(file_name)
    print(f"Generated Excel file: {file_name}")


get_campaign_members(get_campaign(auth_token))
df2 = get_campaign_members(campaign_id)
fundraising(df2)
format_data()
